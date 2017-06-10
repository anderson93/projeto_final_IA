'''
    File name: comparator.py
    Author: Anderson Henrique de Oliveira Conceicao
    Date created: 28/05/2017
    Date last modified: 02/06/2017
    Python Version: 2.7
'''
import pso as PSO
from math import sqrt
from openpyxl import Workbook, workbook, load_workbook

#Opening the data sheet==============================================================================================================================
wb = load_workbook('v_final_FESP_PadroesDeInfracao.xlsx', read_only = True, data_only=True)
ws = wb.active          #Grab the active worksheet

#Acquiring the optimized parameters=================================================================================================================
pso_COST, pso_POSITION = PSO.run() 

#Compare function===================================================================================================================================
def compare(ROW_AUX, pso_POSITION):
    
    #Get the values of the table====================================================================================================================
    AAET_SHEET = ws.cell(row=ROW_AUX,column=70).value        
    AAST_SHEET = ws.cell(row=ROW_AUX,column=75).value
    VI_SHEET   = ws.cell(row=ROW_AUX,column=81).value
    QMA_SHEET  = ws.cell(row=ROW_AUX,column=82).value
    IMAI_SHEET = ws.cell(row=ROW_AUX,column=83).value
    VA_SHEET   = ws.cell(row=ROW_AUX,column=80).value

    ALVO_SHEET = ws.cell(row=ROW_AUX,column=14).value
    
    #Fetching the parameters=======================================================================================================================
    AAET = pso_POSITION[0]      # x = AAET = x[0] ;
    AAST = pso_POSITION[1]      # y = AAST = x[1] ;
    VI   = pso_POSITION[2]      # z = VI = x[2] ;
    QMA  = pso_POSITION[3]      # v = QMA = x[3] ;
    IMAI = pso_POSITION[4]      # w = IMAI = x[4] ;
    VA   = pso_POSITION[5]      # u = VA = x[5].
    
    print "\n", "These are the selected parameters by the PSO:"
    print "CA =", "{:.4f}".format((VA/(AAST-AAET))**(-1)) ,\
          "CP =", "{:.4f}".format(VI/(AAST-AAET)), "\n" , \
          "CC =", "{:.4f}".format((AAST/AAET)**(-1)), \
          "CI =", "{:.4f}".format(QMA/IMAI), "\n"

    #Calculating the optimized parameters===========================================================================================================
    #Calculating the CA coefficient
    try: CA = (VA/(AAST-AAET))**(-1)      
    except ZeroDivisionError: CA = 'inf'
    #Calculating the CP coefficient
    try: CP = VI/(AAST-AAET)
    except ZeroDivisionError: CP = 'inf'
    #Calculating the CC coefficient
    try: CC = (AAST/AAET)**(-1)
    except ZeroDivisionError: CC = 'inf'
    #Calculating the CI coefficient
    try: CI = QMA/IMAI
    except ZeroDivisionError: CI = 'inf'
    
    #Calculating the parameters from the sheet======================================================================================================
    #Calculating the CA coefficient from the sheet
    try: CA_SHEET = VA_SHEET/(AAST_SHEET-AAET_SHEET)
    except ZeroDivisionError: CA_SHEET = 'inf'
    #Calculating the CP coefficient from the sheet
    try: CP_SHEET = VI_SHEET/(AAST_SHEET-AAET_SHEET)
    except ZeroDivisionError: CP_SHEET = 'inf'    
    #Calculating the CC coefficient from the sheet
    try: CC_SHEET = AAST_SHEET/AAET_SHEET        
    except ZeroDivisionError: CC_SHEET = 'inf'    
    #Calculating the CI coefficient from the sheet
    try: CI_SHEET = QMA_SHEET/IMAI_SHEET    
    except ZeroDivisionError: CI_SHEET = 'inf'
    
    #Printing the parameters========================================================================================================================
    print "Those are the selected contributor's parameters:"
    print "CA_SHEET =", "{:.4f}".format(CA_SHEET) ,"CP_SHEET =", "{:.4f}".format(CP_SHEET),\
    "\n", "CC_SHEET =", "{:.4f}".format(CC_SHEET), " CI_SHEET =", "{:.4f}".format(CI_SHEET), "\n"
    
    print "These are the selected parameters by the PSO:"
    print "AAET =", "{:.4f}".format(AAET), "AAST =", "{:.4f}".format(AAST), "VI =", "{:.4f}".format(VI),\
    "\n", " QMA =", "{:.4f}".format(QMA),  "IMAI =", "{:.4f}".format(IMAI), "VA =", "{:.4f}".format(VA), "\n"
    
    print "Those are the selected contributor's parameters:"
    print "AAET_SHEET =", "{:.4f}".format(AAET_SHEET), "AAST_SHEET = ", "{:.4f}".format(AAST_SHEET), "VI_SHEET =", "{:.4f}".format(VI_SHEET),\
    "\n", " QMA_SHEET =", "{:.4f}".format(QMA_SHEET),  "IMAI_SHEET =", "{:.4f}".format(IMAI_SHEET), " VA_SHEET =", "{:.4f}".format(VA_SHEET), "\n"

    #Measuring the Euclidian distance==============================================================================================================
    distance = sqrt((AAET - AAET_SHEET)**2 + (AAST - AAST_SHEET)**2\
                   +(VA - VA_SHEET)**2 + (VI - VI_SHEET)**2\
                   +(QMA - QMA_SHEET)**2 + (IMAI - IMAI_SHEET)**2)
    
    return distance, ALVO_SHEET
#Calculating the euclidean distance from the optimized parameters to the selected contributor======================================================
choose_row = input("\n Choose a row between:")
euclidian_dist, ALVO = compare(choose_row, pso_POSITION)
#Printing the distance=============================================================================================================================
print "The Euclidean distance between the chosen contributor and the ", "\n", \
"optimal profile defined by the PSO algorithm is:", "{:.6f}".format(euclidian_dist), "\n"
print "The chosen row is supposed to be:", ALVO




